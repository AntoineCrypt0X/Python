import tweepy
import time
import datetime
import random
from openpyxl import load_workbook

api_key="your api_key"
api_key_secret="your api_key_secret"
bearer="your bearer"
access_token="your access_token"
access_token_secret="your access_token_secret"

autherticator= tweepy.OAuthHandler(api_key, api_key_secret)
autherticator.set_access_token(access_token, access_token_secret)
api = tweepy.API(autherticator)

# Like / Retweet the last tweet of a list of twitter Accounts
#The Twitter API is subject to rate limits per minute and per hour. Check the latest updates to adjust the waiting times between each request.
def botLikeRetweetInfluencer():

    chemin = "yourExcelPath.xlsx"

    #parameters
    delaybetweenAction=1800

    #Load the sheet named "Data" of the xlsx
    wb = load_workbook(chemin)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    ListAccount=[]

    #Load the list of accounts in the Excel file, column A, start line 2.
    i=2
    while str(sheet['A' + str(i)].value) != "None":
        ListAccount.append(sheet['A' + str(i)].value)
        i=i+1
    print(ListAccount)

    countlike=0
    countTweet=0
    NbLoop=0

    while True:

        print(f"\n{datetime.datetime.utcnow()}\n")
        NbLoop=NbLoop+1

        for username in ListAccount:

            countTweet = countTweet + 1

            try:
                
                user = api.get_user(screen_name=username)
                print('Account :' + username)

                # Get the last tweet of the account
                tweet_user = api.user_timeline(user_id=user.id, count=1, exclude_replies=False, include_rts=True)

                for tweet in tweet_user:

                    print('Tweet text :' + tweet.text)

                    try:
                        api.create_favorite(tweet.id)
                        # api.retweet(tweet.id) 
                        countlike = countlike + 1
                    except Exception as e1:
                        print('error', e1)
                        # return e1
                        print("Already Liked or Retweet")
                    else:
                        a = 1

                    time.sleep(delaybetweenAction)

            except Exception as e2:
                print('error', e2)
                # return e2
            else:
                a = 2

            countTweet = countTweet + 1
            print("Ratio Tweet Liked: " + str(countlike) + "/" + str(countTweet))
            print("\n")

        print("Loop number :" + str(NbLoop))

botLikeRetweetInfluencer()
