import requests
import telebot
from openpyxl import load_workbook
import datetime
import time

#API BOT
API_KEY = "YourBotTelegramAPIkey"
bot = telebot.TeleBot(API_KEY, parse_mode=None)
#https://api.telegram.org/botYOURAPIKEY/getUpdates

#BOT that posts every 30 minutes a messsage from a list of text saved in an Excel file
def BOT_post_message(chatId):

    delaybetweenAction=1800
    pathExcel = "yourExcelPath.xlsx" #Excel that contains the text of the comments

    #Load the sheet named "Data" of the xlsx
    wb = load_workbook(pathExcel)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    ListMessage = []

    #Load the list of tweet text in the Excel file, column A, start line 2.
    j = 2
    while str(sheet['A' + str(j)].value) != "None":
        ListMessage.append(sheet['A' + str(j)].value)
        j = j + 1
    print(ListMessage)

    NbMessage = len(ListMessage)

    for indexMessage in range(0, NbMessage):

        print(f"\n{datetime.datetime.utcnow()}\n")
        print("index Message : " + str(indexMessage))

        try:
            post=ListMessage[indexMessage]
            bot.send_message(chat_id=chatId, text=post)
            print("Telegram post created")

        except Exception as e1:
            print('error1', e1)

        else:
             a = 1

        time.sleep(delaybetweenAction)



