import tweepy
import time
import datetime
from openpyxl import load_workbook

api_key="your api_key"
api_key_secret="your api_key_secret"
bearer="your bearer"
access_token="your access_token"
access_token_secret="your access_token_secret"

autherticator= tweepy.OAuthHandler(api_key, api_key_secret)
autherticator.set_access_token(access_token, access_token_secret)
api = tweepy.API(autherticator)

#BOT that create a post every 30 minutes from a list of text saved in an Excel file
def tweeter():

    delaybetweenAction=1800
    pathExcel = "yourExcelPath.xlsx" #Excel that contains the text of the comments

    #Load the sheet named "Data" of the xlsx
    wb = load_workbook(pathExcel)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    ListMessage = []

    #Load the list of tweet text in the Excel file, column A, start line 2.
    j = 2
    while str(sheet['A' + str(j)].value) != "None":
        ListMessage.append(sheet['A' + str(j)].value)
        j = j + 1
    print(ListMessage)

    Nbtweet = len(ListMessage)

    for indexMessage in range(0, Nbtweet):

        print(f"\n{datetime.datetime.utcnow()}\n")
        print("index Message : " + str(indexMessage))

        try:
            _post = ListMessage[indexMessage]
            print(_post)
            api.update_status(status=_post)
            print("Tweet created")

        except Exception as e1:
            print('error', e1)
            # return e1
            print("Already tweeted or error")

        time.sleep(delaybetweenAction)

tweeter()

