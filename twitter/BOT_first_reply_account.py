import tweepy
import time
from openpyxl import load_workbook

api_key="your api_key"
api_key_secret="your api_key_secret"
bearer="your bearer"
access_token="your access_token"
access_token_secret="your access_token_secret"

autherticator= tweepy.OAuthHandler(api_key, api_key_secret)
autherticator.set_access_token(access_token, access_token_secret)
api = tweepy.API(autherticator)

# Be the first comment of a tweet of a account. Get the last tweet of a account and check if the BOT has already reply or not  
#The Twitter API is subject to rate limits per minute and per hour. Check the latest updates to adjust the waiting times between each request.
def firstReply(username):

    delaybetweenRequest = 1
    delaybetweenAction=1800
    chemin = "yourExcelPath.xlsx" #Excel that contains the text of the comments
    
    #Load the sheet named "Data" of the xlsx
    wb = load_workbook(chemin)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    ListComment = []
    replied_tweet_id=[]

    #Load the list of comments in the Excel file, column A, start line 2.
    j = 2
    while str(sheet['A' + str(j)].value) != "None":
        ListComment.append(sheet['A' + str(j)].value)
        j = j + 1
    print(ListComment)

    countReply = 0
    indexComment = 0
    NbComment = len(ListComment)

    while True:

        try:
            user = api.get_user(screen_name=username)
            #get the last tweet
            lasttweet = api.user_timeline(user_id=user.id, count=1, exclude_replies=False, include_rts=True)

            for tweet in lasttweet:
                # Check if the BOT has already replied
                if not tweet.id in replied_tweet_id:

                    reponse = ListComment[indexComment]
                    print("Text :" + tweet.text)
                    post = api.update_status(status=reponse, in_reply_to_status_id=tweet.id, auto_populate_reply_metadata=True)
                    # The tweet Id is saved to not reply a second time
                    replied_tweet_id.append(tweet.id)
                    countReply+=1
                    indexComment+=1
                    q, indexComment = divmod(indexComment, NbComment)
                    print("Tweet Replied : " + str(countReply))
                    time.sleep(delaybetweenAction)

        except Exception as e:
            print('error', e)
            # return e
            print("Already replied or error")

        time.sleep(delaybetweenRequest)

firstReply("crypto_futur")
