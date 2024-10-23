import tweepy
import time
import datetime
from openpyxl import load_workbook

api_key="your api_key"
api_key_secret="your api_key_secret"
bearer="your bearer"
access_token="your access_token"
access_token_secret="your access_token_secret"

autherticator= tweepy.OAuthHandler(api_key, api_key_secret)
autherticator.set_access_token(access_token, access_token_secret)
api = tweepy.API(autherticator)

#BOT that create a post with an image and a text every 30 minutes, from a list of .png and text saved in an Excel file
def tweeter():

    delaybetweenAction=1800
    pathExcel = "yourExcelPath.xlsx" #Excel that contains the text of the comments

    #Load the sheet named "Data" of the xlsx
    wb = load_workbook(pathExcel)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    ListImage = []
    ListMessage =[]

    #Load the list of image (ex image1.png) and text in the Excel file, column A and B, start line 2.
    j = 2
    while str(sheet['A' + str(j)].value) != "None":
        ListImage.append(sheet['A' + str(j)].value)
        ListMessage.append(sheet['B' + str(j)].value)
        j = j + 1
    print(ListImage)
    print(ListMessage)

    NbImage = len(ListImage)

    for indexImage in range(0, NbImage):

        print(f"\n{datetime.datetime.utcnow()}\n")
        print("index Message : " + str(indexImage))

        try:
            image = ListImage[indexImage]
            message = ListMessage[indexImage]

            _media = api.media_upload(image)
            print(str(image),message)
            api.update_status(status=message, media_ids=[_media.media_id])
            print("Tweet created")

        except Exception as e1:
            print('error', e1)
            # return e1
            print("Already tweeted or error")

        time.sleep(delaybetweenAction)

tweeter()




