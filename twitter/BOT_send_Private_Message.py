import tweepy
import time
import datetime
from openpyxl import load_workbook

api_key="your api_key"
api_key_secret="your api_key_secret"
bearer="your bearer"
access_token="your access_token"
access_token_secret="your access_token_secret"

autherticator= tweepy.OAuthHandler(api_key, api_key_secret)
autherticator.set_access_token(access_token, access_token_secret)
api = tweepy.API(autherticator)


# send private message to a list of account that has been extracted
#The Twitter API is subject to rate limits per minute and per hour. Check the latest updates to adjust the waiting times between each request.
def sendlist():

    print(f"\n{datetime.datetime.now()}\n")

    message= "Your Message"

    #Load the sheet named "Data" of the xlsx
    wb = load_workbook(chemin)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    ListAccount=[]

    #Load the list of accounts in the Excel file, column A, start line 2.
    i=2
    while str(sheet['A' + str(i)].value) != "None":
        ListAccount.append(sheet['A' + str(i)].value)
        i=i+1
    print(ListAccount)

    countSend = 0
    index=0

    for username in ListAccount:

        try:
            user = api.get_user(screen_name=username)
            introduction="Hello " + user.name + ", \n\n\n"
            _message=introduction + message
            api.send_direct_message(user.id, _message)
            countSend = countSend + 1
            index = index + 1

            print("Message send to: " + username + ", total send :" + str(countSend) + "/" + str(index))
        except Exception as e1:
            print('error', e1)
            # return e1
        finally:
            time.sleep(30)
            
sendlist()
