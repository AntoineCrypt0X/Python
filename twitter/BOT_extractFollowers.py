import tweepy
import time
from openpyxl import load_workbook


api_key="your api_key"
api_key_secret="your api_key_secret"
bearer="your bearer"
access_token="your access_token"
access_token_secret="your access_token_secret"

autherticator= tweepy.OAuthHandler(api_key, api_key_secret)
autherticator.set_access_token(access_token, access_token_secret)
api = tweepy.API(autherticator)

# extract the followers from an account. 
#The Twitter API is subject to rate limits per minute and per hour. Check the latest updates to adjust the waiting times between each request.
def botExtractFollowers(username):

    #we load the Excel if it exists(in case the BOT has stopped), else we create it.
    try:
        path_excel="yourExcelPath/followers_" + username + ".xlsx"
        wb = load_workbook(path_excel)
        index=0
        for index in range(len(wb.sheetnames)):
            if wb.sheetnames[index] == "Followers":
                break
        wb.active = index
        sheet = wb.active
    except Exception as e:
        print(e)
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Followers"
        sheet['A1'] = "cursor"
        sheet['B1'] = "line"
        sheet['B3'] = "column"
        sheet['C1'] = "screen_name"
        wb.save(path_excel)
    else:
        a = 1
     
     #we load the last cell filled and the last follower page treated (cursor)
    _index = sheet['B2'].value
    if _index is None:
        _index=2
    _column = sheet['B4'].value
    if _column is None:
        _column=3
    _cursor=sheet['A2'].value
    if _cursor is None:
        #first page of followers
        _cursor=-1
    
    followersNb=(_index-2) + 1000000*(_column-3)
    user = api.get_user(screen_name=username)
    totalFollowers = dict(user._json)["followers_count"]
    
     # last page of followers is cursor=0
    while _cursor!=0:

        try:

            list_follower=user.followers(cursor=curseur, count=200)
            next_cursor=list_follower[1][1]

            list_temp=list_follower[0]

            for follower in list_temp:
                #When the line 1000000 is reached, we write in the column next to, and restart line 2
                if _index>1000001:
                    _column=_column+1
                    _index=2

                _cell = sheet.cell(row=_index, column=_column)
                _cell.value = follower.screen_name
                _index = _index + 1
                followersNb=followersNb+1
            print("Followers :" + str(followersNb) + "/" + str(totalFollowers))
            curseur = next_cursor
            time.sleep(30)

        except Exception as e1:
            print('error', e1)
            # return e1


botExtractFollowers("Bitcoin_archive")
